"""ElstatProvider — Hellenic Statistical Authority (V2 stateless).

Dispatcher ruft fetch_series(spec) pro data_series-Row. KEINE indicator/country/
source-Knowledge, keine Loops ueber Configs.

ELSTAT publiziert keine SDMX/JSON-API; Time-Series kommen als XLS/XLSX-Dateien
ueber das Liferay-Publikationsportal (www.statistics.gr). Stabile numerische
documentID je Veroeffentlichung (publication code wie DKT87, SJO02, ...).

SeriesSpec-Konventionen:
  - spec.series_id ist EINE der folgenden Formen:
      a) "ELSTAT/<publication>/<doc_id>"
           -> einzelne Standard-Reihe (CPI/IPI/PPI/UNEMP/RETAIL/GDP/EMPLOYED)
      b) "ELSTAT/DKT87/114839/<group_code>"
           -> CPI sub-indices (Tabelle VI). group_code: G01..G13.
      c) "ELSTAT/SFC02/115720"  + extra_params.column in {imports,exports,balance}
           -> Handel: drei Spalten teilen sich denselben Workbook.
  - spec.extra_params optional:
      {'column': 'imports' | 'exports' | 'balance'}  (nur SFC02)
      {'group': 'G01'..'G13'}  (alternativ zur Pfad-Variante b)
  - spec.freq_hint steuert die Datumsnormalisierung (M / Q).

Bekannte Series-IDs (Smoke-Tests):
  ELSTAT/DKT87/114838                  -> CPI 1959-… (M, NSA, base 2020=100)
  ELSTAT/DKT21/114474                  -> Industrial Production SA (M, base 2021=100)
  ELSTAT/SJO02/116021                  -> Unemployment SA (M, %)
  ELSTAT/DKT15/587776                  -> Total PPI (M, base 2021=100, rolling)
  ELSTAT/DKT39/500036                  -> Retail Turnover SA (M, base 2021=100)
  ELSTAT/SFC02/115720 + column=imports -> Imports (M, mEUR)
  ELSTAT/SFC02/115720 + column=exports -> Exports (M, mEUR)
  ELSTAT/SFC02/115720 + column=balance -> Trade Balance (M, mEUR)
  ELSTAT/SEL84/115384                  -> Real GDP SA chain-linked (Q, mEUR)
  ELSTAT/SJO01/115983                  -> Employed persons (Q, thousand)
  ELSTAT/DKT87/114839/G01              -> CPI Food sub-index (M)
  ELSTAT/DKT87/114839/G03              -> CPI Clothing sub-index (M)
  ELSTAT/DKT87/114839/G04              -> CPI Housing/Utilities sub-index (M)
  ELSTAT/DKT87/114839/G07              -> CPI Transport sub-index (M)
  ELSTAT/DKT87/114839/G09              -> CPI Recreation sub-index (M)
  ELSTAT/DKT87/114839/G10              -> CPI Education sub-index (M)
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
import requests
import xlrd

from pipeline.base_provider import (
    BaseProvider, SeriesSpec, Observation,
    ProviderError, TransientProviderError,
)
from pipeline.transforms import normalize_date
from pipeline.dispatcher import register_provider

HDR = {
    "User-Agent": "EconPulse/0.1 (Sebastian/SVM-AG)",
    "Accept": "*/*",
}

MONTHS_EN = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

MONTHS_ROMAN = {
    "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6,
    "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12,
}

# CPI Tabelle VI sub-index — Gruppe-Code -> COICOP-Label-Prefix (lowercase startswith)
CPI_SUBGROUP_PREFIX: dict[str, str] = {
    "G01": "food and non-alcoholic",
    "G03": "clothing and footwear",
    "G04": "housing, water, electricity",
    "G07": "transport",
    "G09": "recreation, sport and culture",
    "G10": "education services",
}


# ---------------- HTTP ----------------

def _resource_url(document_id: int) -> str:
    """Stable Liferay 'downloadResources' URL fuer ein ELSTAT-Dokument."""
    portlet = "documents_WAR_publicationsportlet_INSTANCE_Mr0GiQJSgPHd"
    return (
        "https://www.statistics.gr/en/statistics?"
        f"p_p_id={portlet}"
        "&p_p_lifecycle=2&p_p_state=normal&p_p_mode=view&p_p_cacheability=cacheLevelPage"
        f"&_{portlet}_javax.faces.resource=document"
        f"&_{portlet}_ln=downloadResources"
        f"&_{portlet}_documentID={document_id}"
        f"&_{portlet}_locale=en"
    )


def _http_get(url: str) -> bytes:
    try:
        r = requests.get(url, headers=HDR, timeout=60)
    except (requests.ConnectionError, requests.Timeout) as e:
        raise TransientProviderError(f"elstat network: {e}") from e
    if r.status_code >= 500:
        raise TransientProviderError(f"elstat HTTP {r.status_code}")
    if r.status_code == 404:
        raise ProviderError(f"elstat HTTP 404: {url}")
    if r.status_code >= 400:
        raise ProviderError(f"elstat HTTP {r.status_code}: {r.text[:200]}")
    return r.content


# ---------------- Per-Publication-Parser ----------------

def _parse_cpi(content: bytes) -> list[tuple[date, float]]:
    """DKT87 Table IV — Monthly CPI matrix (XLSX)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    out: list[tuple[date, float]] = []
    current_years: list[tuple[int, int]] = []
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if "Month" in cells:
            idx = cells.index("Month")
            current_years = [
                (j, c) for j, c in enumerate(cells[idx + 1:], start=idx + 1)
                if isinstance(c, int) and 1900 <= c <= 2100
            ]
            continue
        if not current_years:
            continue
        month_idx = next((i for i, c in enumerate(cells) if c in MONTHS_EN), None)
        if month_idx is None:
            continue
        month_num = MONTHS_EN.index(cells[month_idx]) + 1
        for col_j, year in current_years:
            v = cells[col_j] if col_j < len(cells) else None
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                try:
                    out.append((date(year, month_num, 1), float(v)))
                except ValueError:
                    pass
    out.sort()
    return out


def _parse_ipi_sa(content: bytes) -> list[tuple[date, float]]:
    """DKT21 Industrial Production SA (XLSX)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = (wb["INDUSTRIAL PRODUCTION INDEX"]
          if "INDUSTRIAL PRODUCTION INDEX" in wb.sheetnames else wb.active)
    out: list[tuple[date, float]] = []
    current_year: int | None = None
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not cells:
            continue
        c0 = cells[0]
        if isinstance(c0, int) and 1900 <= c0 <= 2100:
            current_year = c0
        if current_year is None or len(cells) < 3:
            continue
        c1 = cells[1] if len(cells) > 1 else None
        if not isinstance(c1, int) or not (1 <= c1 <= 12):
            continue
        c2 = cells[2]
        if not isinstance(c2, (int, float)) or isinstance(c2, bool):
            continue
        out.append((date(current_year, c1, 1), float(c2)))
    out.sort()
    return out


def _parse_unemployment_sa(content: bytes) -> list[tuple[date, float]]:
    """SJO02 Tabelle 1A — Unemployment Rate SA (XLS legacy). Col 8 = SA rate."""
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    out: list[tuple[date, float]] = []
    current_year: int | None = None
    for r_i in range(ws.nrows):
        row = [ws.cell_value(r_i, c) for c in range(ws.ncols)]
        c0 = row[0]
        if isinstance(c0, float) and 1900 <= c0 <= 2100 and c0.is_integer():
            current_year = int(c0)
            continue
        if isinstance(c0, str) and c0 in MONTHS_EN and current_year is not None:
            month_num = MONTHS_EN.index(c0) + 1
            if len(row) > 8 and isinstance(row[8], (int, float)) and row[8] != "":
                out.append((date(current_year, month_num, 1), float(row[8])))
    out.sort()
    return out


def _parse_ppi(content: bytes) -> list[tuple[date, float]]:
    """DKT15 'Total PPI' (XLS legacy) — Code 0020 Overall Market row."""
    wb = xlrd.open_workbook(file_contents=content)
    ws = (wb.sheet_by_name("Total PPI") if "Total PPI" in wb.sheet_names()
          else wb.sheet_by_index(0))
    header_row: int | None = None
    period_cols: list[tuple[int, date]] = []
    for r_i in range(min(ws.nrows, 30)):
        row = [ws.cell_value(r_i, c) for c in range(ws.ncols)]
        labels = [(j, c) for j, c in enumerate(row)
                  if isinstance(c, str) and len(c) == 7 and c[4] == "_"]
        if len(labels) >= 2:
            header_row = r_i
            for j, lbl in labels:
                try:
                    yr, mo = int(lbl[:4]), int(lbl[5:7])
                    if 1900 <= yr <= 2100 and 1 <= mo <= 12:
                        period_cols.append((j, date(yr, mo, 1)))
                except ValueError:
                    continue
            break
    if header_row is None or not period_cols:
        return []
    out: list[tuple[date, float]] = []
    for r_i in range(header_row + 1, ws.nrows):
        c0 = ws.cell_value(r_i, 0)
        if str(c0).strip() in ("0020",) or (isinstance(c0, float) and int(c0) == 20):
            for col_j, dt in period_cols:
                v = ws.cell_value(r_i, col_j)
                if isinstance(v, (int, float)) and v != "" and not isinstance(v, bool):
                    out.append((dt, float(v)))
            break
    out.sort()
    return out


def _parse_retail_sa(content: bytes) -> list[tuple[date, float]]:
    """DKT39 SA Turnover (XLSX) — sheet 'TABLE 3', col 1 = Overall Index."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    out: list[tuple[date, float]] = []
    current_year: int | None = None
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not cells:
            continue
        c0 = cells[0]
        if not isinstance(c0, str):
            continue
        label = c0.strip()
        parts = label.split()
        month_token: str | None = None
        if len(parts) >= 2 and parts[0].isdigit() and 1900 <= int(parts[0]) <= 2100:
            current_year = int(parts[0])
            month_token = parts[1]
        elif label in MONTHS_ROMAN:
            month_token = label
        if month_token is None or month_token not in MONTHS_ROMAN or current_year is None:
            continue
        v = cells[1] if len(cells) > 1 else None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            out.append((date(current_year, MONTHS_ROMAN[month_token], 1), float(v)))
    out.sort()
    return out


def _parse_trade(content: bytes) -> list[tuple[date, float, float, float]]:
    """SFC02 — Trade Balance SDDS sheet. Returns (date, imports, exports, balance)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb["TRADE BALANCE SDDS MONTHLY DATA"]
    out: list[tuple[date, float, float, float]] = []
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if len(cells) < 11:
            continue
        yr, mo = cells[0], cells[1]
        if not (isinstance(yr, int) and isinstance(mo, int)
                and 2000 <= yr <= 2100 and 1 <= mo <= 12):
            continue
        imp, exp, bal = cells[2], cells[7], cells[10]
        if not all(isinstance(v, (int, float)) and not isinstance(v, bool)
                   for v in (imp, exp, bal)):
            continue
        out.append((date(yr, mo, 1), float(imp), float(exp), float(bal)))
    out.sort()
    return out


def _parse_gdp(content: bytes) -> list[tuple[date, float]]:
    """SEL84 — Quarterly GDP SA chain-linked (XLS legacy, wide-transposed)."""
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    header_row: int | None = None
    period_cols: list[tuple[int, date]] = []
    for r_i in range(min(ws.nrows, 15)):
        row = [ws.cell_value(r_i, c) for c in range(ws.ncols)]
        labels = []
        for j, c in enumerate(row):
            if isinstance(c, str) and " Q" in c:
                parts = c.replace("\xa0", " ").split()
                if (len(parts) == 2 and parts[0].isdigit()
                        and parts[1] in ("Q1", "Q2", "Q3", "Q4")):
                    yr = int(parts[0])
                    q = int(parts[1][1])
                    labels.append((j, date(yr, q * 3, 1)))
        if len(labels) >= 4:
            header_row = r_i
            period_cols = labels
            break
    if header_row is None:
        return []
    out: list[tuple[date, float]] = []
    for r_i in range(header_row + 1, ws.nrows):
        c0 = ws.cell_value(r_i, 0)
        if isinstance(c0, str) and "Gross Domestic Product" in c0:
            for col_j, dt in period_cols:
                v = ws.cell_value(r_i, col_j)
                if isinstance(v, (int, float)) and v != "" and not isinstance(v, bool):
                    out.append((dt, float(v)))
            break
    out.sort()
    return out


def _parse_employed(content: bytes) -> list[tuple[date, float]]:
    """SJO01 Table 3 — Persons employed quarterly (XLS legacy).
    Bevorzugt NACE Rev 2 Block (spaeter in der Datei -> ueberschreibt Rev 1).
    """
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    rows = [[ws.cell_value(r_i, c) for c in range(ws.ncols)] for r_i in range(ws.nrows)]

    def parse_q_label(s) -> date | None:
        if not isinstance(s, str):
            return None
        s2 = " ".join(s.split())
        for prefix, q in (("1st", 1), ("2d", 2), ("3d", 3), ("4th", 4)):
            if s2.lower().startswith(prefix):
                for t in s2.split():
                    if t.isdigit() and 1900 <= int(t) <= 2100:
                        return date(int(t), q * 3, 1)
        return None

    out: dict[date, float] = {}
    last_header_cols: list[tuple[int, date]] = []
    for row in rows:
        labels = [(j, parse_q_label(c)) for j, c in enumerate(row)]
        labels = [(j, d) for j, d in labels if d is not None]
        if len(labels) >= 4:
            last_header_cols = labels
            continue
        col1 = row[1] if len(row) > 1 else None
        if isinstance(col1, str) and "Total employed" in col1 and last_header_cols:
            for col_j, dt in last_header_cols:
                v = row[col_j] if col_j < len(row) else None
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    out[dt] = float(v)
    return sorted(out.items())


def _parse_cpi_subgroup(content: bytes, group_code: str) -> list[tuple[date, float]]:
    """DKT87 Table VI — COICOP sub-index fuer eine Gruppe (XLSX).
    group_code: 'G01'..'G13' (siehe CPI_SUBGROUP_PREFIX). Match via Label-Prefix.
    """
    prefix = CPI_SUBGROUP_PREFIX.get(group_code.upper())
    if not prefix:
        raise ProviderError(f"elstat: unknown CPI sub-group '{group_code}'")
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    out: list[tuple[date, float]] = []
    current_year: int | None = None
    in_block = False
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not cells:
            continue
        c0 = cells[0]
        if not isinstance(c0, str):
            continue
        s = c0.strip()
        if s.lower().startswith("year"):
            for p in s.split():
                if p.isdigit() and 1900 <= int(p) <= 2100:
                    current_year = int(p)
                    in_block = False
                    break
            continue
        if s.lower().startswith("groups of") and current_year is not None:
            in_block = True
            continue
        if in_block and current_year is not None:
            label = s
            if label and label[0].isdigit():
                label = label.split(maxsplit=1)[-1] if " " in label else label
            if label.lower().startswith(prefix):
                for m_idx in range(1, 13):
                    v = cells[m_idx] if m_idx < len(cells) else None
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        out.append((date(current_year, m_idx, 1), float(v)))
    out.sort()
    return out


# ---------------- series_id-Routing ----------------

def _parse_series_id(series_id: str) -> tuple[str, int, str | None]:
    """ELSTAT-series_id -> (publication, doc_id, suffix).

    Akzeptiert:
      'ELSTAT/<PUB>/<DOC_ID>'
      'ELSTAT/<PUB>/<DOC_ID>/<SUFFIX>'   (CPI sub-indices: SUFFIX = G01..G13)
    """
    sid = (series_id or "").strip()
    if not sid:
        raise ProviderError("elstat: empty series_id")
    if sid.upper().startswith("ELSTAT/"):
        sid = sid.split("/", 1)[1]
    parts = sid.split("/")
    if len(parts) < 2:
        raise ProviderError(f"elstat: cannot parse series_id '{series_id}'")
    pub = parts[0]
    try:
        doc_id = int(parts[1])
    except ValueError as e:
        raise ProviderError(
            f"elstat: series_id doc_id not int in '{series_id}'") from e
    suffix = parts[2] if len(parts) >= 3 else None
    return pub.upper(), doc_id, suffix


class ElstatProvider(BaseProvider):
    name = "elstat"
    display_name = "ELSTAT (Hellenic Statistical Authority)"

    def fetch_series(self, spec: SeriesSpec) -> list[Observation]:
        pub, doc_id, suffix = _parse_series_id(spec.series_id)
        ep = spec.extra_params or {}
        freq = spec.freq_hint or "M"
        conv = spec.conversion or 1.0

        content = _http_get(_resource_url(doc_id))

        pairs: list[tuple[date, float]]
        if pub == "DKT87" and doc_id == 114838:
            pairs = _parse_cpi(content)
        elif pub == "DKT87" and doc_id == 114839:
            group = (ep.get("group") or suffix or "").upper()
            if not group:
                raise ProviderError(
                    "elstat: CPI sub-index needs group_code "
                    "(series_id suffix or extra_params.group)"
                )
            pairs = _parse_cpi_subgroup(content, group)
        elif pub == "DKT21":
            pairs = _parse_ipi_sa(content)
        elif pub == "SJO02":
            pairs = _parse_unemployment_sa(content)
        elif pub == "DKT15":
            pairs = _parse_ppi(content)
        elif pub == "DKT39":
            pairs = _parse_retail_sa(content)
        elif pub == "SFC02":
            col = ((ep.get("column") or suffix) or "").lower()
            if col in ("imp", "imports"):
                pairs = [(d, imp) for d, imp, _, _ in _parse_trade(content)]
            elif col in ("exp", "exports"):
                pairs = [(d, exp) for d, _, exp, _ in _parse_trade(content)]
            elif col in ("bal", "balance", "trade-balance"):
                pairs = [(d, bal) for d, _, _, bal in _parse_trade(content)]
            else:
                raise ProviderError(
                    "elstat: SFC02 trade requires extra_params.column "
                    "in {imports, exports, balance}"
                )
        elif pub == "SEL84":
            pairs = _parse_gdp(content)
        elif pub == "SJO01":
            pairs = _parse_employed(content)
        else:
            raise ProviderError(
                f"elstat: unknown publication '{pub}' (doc_id={doc_id}) in series_id"
            )

        return [
            Observation(
                date=normalize_date(dt, freq),
                value=round(float(v) * conv, 6),
            )
            for dt, v in pairs
        ]


try:
    register_provider(ElstatProvider())
except ProviderError as e:
    print(f"[warn] ElstatProvider not registered: {e}")
