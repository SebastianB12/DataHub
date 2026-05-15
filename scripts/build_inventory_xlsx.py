"""Build a comprehensive Excel inventory of EconPulse data coverage.

Per (country, slug): source vs TE-source match, our latest value vs TE,
YoY computed for index-display indicators, history length, secondary
source presence, etc.

Output: data_inventory.xlsx
Run:    pipeline/.venv/Scripts/python.exe -m scripts.build_inventory_xlsx
"""
from __future__ import annotations

import time
from collections import defaultdict
from datetime import date
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pipeline.db import supabase as sb

OUTPUT = Path("data_inventory.xlsx")

LABELS = {
    "fred": "FRED (Federal Reserve)",
    "eurostat": "Eurostat",
    "ecb": "European Central Bank",
    "ons": "Office for National Statistics (UK)",
    "worldbank": "World Bank",
    "destatis": "Statistisches Bundesamt",
    "bundesbank": "Deutsche Bundesbank",
    "insee": "INSEE",
    "bdf": "Banque de France",
    "ine_es": "INE Spain",
    "istat": "ISTAT (Italy)",
    "cso_ie": "CSO Ireland",
    "inept": "INE Portugal",
    "ine_pt": "INE Portugal",
    "nbb": "National Bank of Belgium",
    "elstat": "ELSTAT (Greece)",
    "scb": "Statistics Sweden (SCB)",
    "scb_se": "Statistics Sweden (SCB)",
    "konj_se": "NIER Konjunkturinstitutet (Sweden)",
    "statpol": "Statistics Poland (GUS)",
    "gus_pl": "Statistics Poland (GUS DBW)",
    "dst": "Statistics Denmark",
    "stat_fi": "Statistics Finland",
    "stat_at": "Statistik Austria",
    "surs_si": "SURS (Slovenia)",
    "csp_lv": "Statistics Latvia (CSP)",
    "insse_ro": "INSSE Romania",
    "stat_ee": "Statistics Estonia",
    "dzs_hr": "DZS Croatian Bureau",
    "statbel": "Statbel (Belgium)",
    "susr_sk": "Štatistický úrad SR (Slovakia)",
    "ksh_hu": "KSH (Hungary)",
    "nso_mt": "NSO Malta",
    "cystat_cy": "Statistical Service of Cyprus",
    "statec_lu": "STATEC (Luxembourg)",
    "eia": "U.S. Energy Information Administration",
    "gacc": "General Administration of Customs of China",
    "akshare": "NBS / PBoC / GACC / SAFE (via akshare)",
    "curated": "Hand-curated",
    "czso": "Czech Statistical Office (ČSÚ)",
    "lsd_lt": "Statistics Lithuania (LSD)",
    "nsi_bg": "NSI Bulgaria (via BNB SDDS Plus)",
    "dbnomics": "DBnomics gateway",
}


def label(code: str) -> str:
    return LABELS.get(code, code)


def retry(fn, retries=5, sleep=1.0):
    for i in range(retries):
        try:
            return fn()
        except Exception:
            if i == retries - 1:
                raise
            time.sleep(sleep * (i + 1))


def fetch_all_paginated(table: str, select: str, **filters):
    rows = []
    page = 0
    while True:
        q = sb.table(table).select(select)
        for k, v in filters.items():
            if isinstance(v, tuple):
                method, arg = v
                q = getattr(q, method)(*arg) if isinstance(arg, tuple) else getattr(q, method)(arg)
            else:
                q = q.eq(k, v)
        q = q.range(page * 1000, page * 1000 + 999)
        r = retry(lambda: q.execute())
        chunk = r.data or []
        rows.extend(chunk)
        if len(chunk) < 1000:
            break
        page += 1
    return rows


def main():
    print("Fetching indicator_sources (active)...")
    src_rows = []
    page = 0
    while True:
        r = retry(lambda: sb.table("indicator_sources")
                  .select("country, indicator, source, series_id, is_default, transform, unit, adjustment, freq_hint, note")
                  .eq("active", True)
                  .range(page * 1000, page * 1000 + 999).execute())
        chunk = r.data or []
        src_rows.extend(chunk)
        if len(chunk) < 1000:
            break
        page += 1
    print(f"  {len(src_rows)} active indicator_sources rows")

    # Group by (country, indicator)
    by_key: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for r in src_rows:
        by_key[(r["country"], r["indicator"])].append(r)

    print("Fetching indicators metadata...")
    indicators = {row["slug"]: row for row in retry(lambda: sb.table("indicators").select("*").execute()).data}

    print("Loading TE inventory snapshots...")
    inv_dir = Path("docs/_te_inventory")
    inventory: dict[tuple[str, str], dict] = {}
    for fp in sorted(inv_dir.glob("*.yaml")):
        cc = fp.stem
        with open(fp, encoding="utf-8") as f:
            d = yaml.safe_load(f) or {}
        for slug, entry in d.items():
            if isinstance(entry, dict):
                inventory[(cc, slug)] = entry

    print("Fetching latest data_points per (country, indicator) — this takes ~1-2 min...")
    # Strategy: get latest 14 points per (country, indicator, source) so we can compute YoY for monthly
    # and first/last date + count. Single bulk fetch is too much — do per-row.
    latest_cache: dict[tuple[str, str, str], list[dict]] = {}
    range_cache: dict[tuple[str, str, str], tuple] = {}

    keys = list(by_key.keys())
    print(f"  {len(keys)} (country, indicator) keys to process")

    rows_for_xlsx = []

    for i, (country, slug) in enumerate(keys):
        if i % 50 == 0:
            print(f"  {i}/{len(keys)}...")

        rows = by_key[(country, slug)]
        default = next((r for r in rows if r["is_default"]), None)
        secondary = [r for r in rows if not r["is_default"]]
        if not default:
            continue

        # latest 14 data points
        try:
            dp = retry(lambda: sb.table("data_points")
                       .select("date, value")
                       .eq("country", country).eq("indicator", slug).eq("source", default["source"])
                       .order("date", desc=True).limit(14).execute()).data
        except Exception:
            dp = []
        # first date + count
        try:
            first = retry(lambda: sb.table("data_points")
                          .select("date")
                          .eq("country", country).eq("indicator", slug).eq("source", default["source"])
                          .order("date", desc=False).limit(1).execute()).data
            cnt = retry(lambda: sb.table("data_points")
                        .select("date", count="exact")
                        .eq("country", country).eq("indicator", slug).eq("source", default["source"])
                        .limit(1).execute()).count
        except Exception:
            first, cnt = [], 0

        latest = dp[0] if dp else None
        latest_val = latest["value"] if latest else None
        latest_date = latest["date"] if latest else None
        first_date = first[0]["date"] if first else None
        n_obs = cnt or 0

        # YoY computation if index/raw display
        ind_meta = indicators.get(slug, {}) or {}
        display = ind_meta.get("default_display") or ind_meta.get("unit_type") or ""
        is_index = "index" in (display or "").lower() or (default.get("unit") or "").lower().startswith("index")
        yoy = None
        if dp and len(dp) >= 13 and latest_val and latest:
            # Find row 12 months back
            lc = {row["date"][:7]: row["value"] for row in dp}
            ld = latest["date"]
            yr, mo, _ = ld.split("-")
            prev_key = f"{int(yr)-1}-{mo}"
            prev = lc.get(prev_key)
            if prev and prev != 0:
                yoy = (latest_val / prev - 1) * 100

        # TE-Inventory lookup
        inv = inventory.get((country, slug), {})
        te_val = inv.get("te_value")
        te_period = inv.get("te_period")
        te_label_str = inv.get("te_label") or ""
        suggested_source = inv.get("suggested_source")
        verified = inv.get("verified", False)

        # Source match
        if verified and suggested_source:
            source_match = default["source"] == suggested_source
            source_match_str = "MATCH" if source_match else "MISMATCH"
        elif verified and not suggested_source:
            source_match_str = "VERIFIED-LICENSED"
        else:
            source_match_str = "UNVERIFIED"

        # Value match — try direct, also compare YoY if both available
        value_match_str = ""
        if te_val is not None and latest_val is not None:
            try:
                te_v = float(te_val)
                # Tolerate within 2%
                diff = abs(latest_val - te_v)
                tol = max(abs(te_v) * 0.02, 0.5)
                if diff <= tol:
                    value_match_str = "MATCH"
                else:
                    # try YoY match if our value is index-level
                    if yoy is not None and abs(yoy - te_v) < 0.5:
                        value_match_str = "MATCH (via YoY)"
                    else:
                        value_match_str = f"DIFF ({latest_val:.2f} vs TE {te_v:.2f})"
            except (TypeError, ValueError):
                value_match_str = "?"

        secondary_str = ", ".join(f"{r['source']}" for r in secondary[:3])

        rows_for_xlsx.append({
            "Country": country,
            "Slug": slug,
            "Indicator Name": ind_meta.get("name", ""),
            "Category": ind_meta.get("category", ""),
            "Tier": ind_meta.get("tier", ""),
            "Default Source": default["source"],
            "Default Source Label": label(default["source"]),
            "TE Source (Inventory)": te_label_str,
            "Source Match": source_match_str,
            "Series ID": default.get("series_id", ""),
            "Default Display": display,
            "Is Index": "YES" if is_index else "",
            "Latest Value (raw)": latest_val,
            "Latest Date": latest_date,
            "YoY % (on-the-fly)": round(yoy, 2) if yoy is not None else None,
            "Unit": default.get("unit", ""),
            "TE Value": te_val,
            "TE Period": te_period,
            "Value Match": value_match_str,
            "History From": first_date,
            "N Observations": n_obs,
            "Secondary Sources": secondary_str,
            "Adjustment": default.get("adjustment", ""),
            "Freq": default.get("freq_hint", ""),
            "Inventory Verified": "YES" if verified else "",
            "Note": (default.get("note") or "")[:200],
        })

    # Sort by country then slug
    rows_for_xlsx.sort(key=lambda r: (r["Country"], r["Slug"]))

    print(f"Writing {len(rows_for_xlsx)} rows to {OUTPUT}...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Coverage"

    headers = list(rows_for_xlsx[0].keys())
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    match_fill = PatternFill("solid", fgColor="C6EFCE")  # green
    mismatch_fill = PatternFill("solid", fgColor="FFC7CE")  # red
    warn_fill = PatternFill("solid", fgColor="FFEB9C")  # yellow
    index_fill = PatternFill("solid", fgColor="DEEBF7")  # light blue

    src_match_col = headers.index("Source Match") + 1
    val_match_col = headers.index("Value Match") + 1
    is_index_col = headers.index("Is Index") + 1
    yoy_col = headers.index("YoY % (on-the-fly)") + 1

    for r_idx, row in enumerate(rows_for_xlsx, start=2):
        for c_idx, h in enumerate(headers, start=1):
            v = row[h]
            ws.cell(row=r_idx, column=c_idx, value=v)
        # Highlight Source Match column
        sm = row["Source Match"]
        if sm == "MATCH":
            ws.cell(row=r_idx, column=src_match_col).fill = match_fill
        elif sm == "MISMATCH":
            ws.cell(row=r_idx, column=src_match_col).fill = mismatch_fill
        elif sm == "VERIFIED-LICENSED":
            ws.cell(row=r_idx, column=src_match_col).fill = warn_fill
        # Highlight Value Match
        vm = row["Value Match"]
        if vm.startswith("MATCH"):
            ws.cell(row=r_idx, column=val_match_col).fill = match_fill
        elif vm.startswith("DIFF"):
            ws.cell(row=r_idx, column=val_match_col).fill = mismatch_fill
        # Highlight Is Index (light blue + bold yoy)
        if row["Is Index"] == "YES":
            ws.cell(row=r_idx, column=is_index_col).fill = index_fill
            ws.cell(row=r_idx, column=yoy_col).fill = index_fill
            ws.cell(row=r_idx, column=yoy_col).font = Font(bold=True)

    # Freeze pane + col widths
    ws.freeze_panes = "C2"
    widths = {
        "Country": 8, "Slug": 30, "Indicator Name": 28, "Category": 12, "Tier": 6,
        "Default Source": 14, "Default Source Label": 28, "TE Source (Inventory)": 28,
        "Source Match": 14, "Series ID": 26, "Default Display": 12, "Is Index": 8,
        "Latest Value (raw)": 14, "Latest Date": 12, "YoY % (on-the-fly)": 14,
        "Unit": 18, "TE Value": 12, "TE Period": 12, "Value Match": 22,
        "History From": 12, "N Observations": 10, "Secondary Sources": 18,
        "Adjustment": 8, "Freq": 6, "Inventory Verified": 10, "Note": 40,
    }
    for c_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(h, 15)
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "EconPulse Data Inventory — Summary"
    ws2["A1"].font = Font(size=14, bold=True)
    total = len(rows_for_xlsx)
    match_src = sum(1 for r in rows_for_xlsx if r["Source Match"] == "MATCH")
    mismatch_src = sum(1 for r in rows_for_xlsx if r["Source Match"] == "MISMATCH")
    unverified = sum(1 for r in rows_for_xlsx if r["Source Match"] == "UNVERIFIED")
    match_val = sum(1 for r in rows_for_xlsx if r["Value Match"].startswith("MATCH"))
    diff_val = sum(1 for r in rows_for_xlsx if r["Value Match"].startswith("DIFF"))
    idx_rows = sum(1 for r in rows_for_xlsx if r["Is Index"] == "YES")
    yoy_vals = sum(1 for r in rows_for_xlsx if r["YoY % (on-the-fly)"] is not None)
    rows_summary = [
        ("Total (country, slug) rows", total),
        ("Source MATCH (verified)", match_src),
        ("Source MISMATCH (verified)", mismatch_src),
        ("Source UNVERIFIED", unverified),
        ("Value MATCH (within 2%)", match_val),
        ("Value DIFF", diff_val),
        ("Index display rows", idx_rows),
        ("YoY computable rows", yoy_vals),
    ]
    for i, (k, v) in enumerate(rows_summary, start=3):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 14

    wb.save(OUTPUT)
    print(f"Done. {OUTPUT} written ({OUTPUT.absolute()})")


if __name__ == "__main__":
    main()
